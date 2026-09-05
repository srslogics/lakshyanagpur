from datetime import date, datetime, timezone
from io import BytesIO

import pytest

from app.models import (
    AttendanceEntry,
    AttendanceRegister,
    BiometricAttendanceDay,
    BiometricImportBatch,
    DeviceAttendanceIdentity,
    Enrollment,
    Student,
    StudentAcademicProfile,
    StaffAttendanceWorkday,
    User,
)
from app.security import create_token, hash_password
from app.importers.biometric_attendance import (
    SheetData,
    parse_essl_form_j_pdf,
    parse_essl_form_j_sheet,
    parse_essl_work_duration_sheet,
)
from app.routers.portal import attendance_rows
from app.routers.attendance import _staff_designation


def setup_students(database):
    operator = User(
        mobile="9000000099",
        full_name="Attendance Operator",
        role="attendance_operator",
        password_hash=hash_password("Password123!"),
    )
    tatva = Student(
        admission_number="LI-2026-20001",
        full_name="Tatva Student",
        mobile="9000000101",
        status="active",
    )
    essential = Student(
        admission_number="LI-2026-20002",
        full_name="Essential Student",
        mobile="9000000102",
        status="active",
    )
    database.add_all([operator, tatva, essential])
    database.flush()
    database.add_all([
        Enrollment(student_id=tatva.id, program="JEE", batch="Tatva", status="active", is_active=True),
        Enrollment(student_id=essential.id, program="Boards", batch="Essential", status="active", is_active=True),
        StudentAcademicProfile(
            student_id=tatva.id,
            source_student_code="T-1",
            batch_name="Tatva",
            source_stream="JEE",
        ),
        StudentAcademicProfile(
            student_id=essential.id,
            source_student_code="E-1",
            batch_name="Essential",
            source_stream="Boards",
        ),
    ])
    database.commit()
    return operator, tatva, essential


def preview(client, headers, content, filename="attendance.csv"):
    response = client.post(
        "/api/attendance/biometric-imports/preview",
        headers=headers,
        files={"file": (filename, content, "text/csv")},
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_biometric_file_import_maps_students_and_keeps_first_daily_punch(client, database):
    operator, tatva, essential = setup_students(database)
    no_punch = Student(
        admission_number="LI-2026-20003",
        full_name="Tatva No Punch",
        mobile="9000000103",
        status="active",
    )
    database.add(no_punch)
    database.flush()
    database.add_all([
        Enrollment(student_id=no_punch.id, program="JEE", batch="Tatva", status="active", is_active=True),
        StudentAcademicProfile(
            student_id=no_punch.id,
            source_student_code="T-2",
            batch_name="Tatva",
            source_stream="JEE",
        ),
    ])
    database.commit()
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    content = (
        b"User ID,Name,Date,Time\n"
        b"101,Tatva Student,12-08-2026,08:15:00\n"
        b"101,Tatva Student,12-08-2026,16:05:00\n"
        b"202,Essential Student,12-08-2026,08:21:00\n"
        b"999,Staff Member,12-08-2026,08:04:00\n"
    )
    staged = preview(client, headers, content)
    assert staged["device"]["serialNumber"] == "ABFR220607313"
    assert staged["sheets"][0]["detected"]["device_id"] == "User ID"

    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "Attendance",
        "deviceIdColumn": "User ID",
        "nameColumn": "Name",
        "datetimeColumn": None,
        "dateColumn": "Date",
        "timeColumn": "Time",
    }
    analysis = client.post(
        "/api/attendance/biometric-imports/analyze",
        headers=headers,
        json=selection,
    )
    assert analysis.status_code == 200, analysis.text
    assert analysis.json()["rowsSeen"] == 4
    assert analysis.json()["uniqueAttendanceDays"] == 3
    assert analysis.json()["duplicateRows"] == 1

    committed = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={
            **selection,
            "mappings": [
                {"deviceUserId": "101", "studentId": tatva.id, "ignore": False},
                {"deviceUserId": "202", "studentId": essential.id, "ignore": False},
                {"deviceUserId": "999", "studentId": None, "ignore": True},
            ],
        },
    )
    assert committed.status_code == 200, committed.text
    body = committed.json()
    assert body["punchesCreated"] == 2
    assert {item["batch"] for item in body["registers"]} == {"Tatva", "Essential"}
    assert all(item["status"] == "submitted" for item in body["registers"])
    assert "published" in body["message"]
    assert database.query(BiometricAttendanceDay).count() == 2
    tatva_day = database.query(BiometricAttendanceDay).filter_by(student_id=tatva.id).one()
    assert tatva_day.first_punch_at.hour == 2
    assert tatva_day.first_punch_at.minute == 45
    assert database.query(BiometricImportBatch).one().rows_seen == 4
    assert database.query(DeviceAttendanceIdentity).filter_by(device_user_id="999").one().is_ignored is True
    registers = database.query(AttendanceRegister).filter_by(register_kind="biometric").all()
    assert len(registers) == 2
    assert all(register.status == "submitted" for register in registers)
    assert database.query(AttendanceEntry).count() == 3
    assert attendance_rows(database, tatva)[0]["status"] == "present"
    assert attendance_rows(database, no_punch)[0]["status"] == "absent"

    bootstrap = client.get("/api/attendance/bootstrap?day=2026-08-12", headers=headers)
    assert bootstrap.status_code == 200
    imported_sessions = [item for item in bootstrap.json()["sessions"] if item["registerKind"] == "biometric"]
    assert {item["batch"] for item in imported_sessions} == {"Tatva", "Essential"}
    tatva_register = next(item for item in imported_sessions if item["batch"] == "Tatva")
    assert tatva_register["registerStatus"] == "submitted"
    assert tatva_register["markedCount"] == tatva_register["studentCount"]
    roster = client.get(f"/api/attendance/manual-registers/{tatva_register['id']}", headers=headers)
    assert roster.status_code == 200
    statuses = {item["studentId"]: item["status"] for item in roster.json()["entries"]}
    assert statuses[tatva.id] == "present"
    assert statuses[no_punch.id] == "absent"


def test_biometric_import_keeps_staff_punches_out_of_student_registers(client, database):
    operator, tatva, _ = setup_students(database)
    staff = User(
        mobile="9000000199",
        full_name="Office Staff",
        role="front_desk",
        password_hash=hash_password("Password123!"),
    )
    database.add(staff)
    database.commit()
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    content = (
        b"User ID,Name,Timestamp\n"
        b"101,Tatva Student,2026-08-12 08:15:00\n"
        b"900,Office Staff,2026-08-12 08:02:00\n"
        b"900,Office Staff,2026-08-12 18:06:00\n"
    )
    staged = preview(client, headers, content)
    assert any(item["id"] == staff.id for item in staged["staff"])
    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "Attendance",
        "deviceIdColumn": "User ID",
        "nameColumn": "Name",
        "datetimeColumn": "Timestamp",
        "dateColumn": None,
        "timeColumn": None,
    }
    analysis = client.post(
        "/api/attendance/biometric-imports/analyze",
        headers=headers,
        json=selection,
    )
    people = {item["deviceUserId"]: item for item in analysis.json()["deviceUsers"]}
    assert people["900"]["staffUserId"] == staff.id
    committed = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={
            **selection,
            "mappings": [
                {"deviceUserId": "101", "studentId": tatva.id, "ignore": False},
                {"deviceUserId": "900", "staffUserId": staff.id, "ignore": False},
            ],
        },
    )
    assert committed.status_code == 200, committed.text
    assert committed.json()["staffPunchesCreated"] == 1
    assert committed.json()["matchedStaff"] == 1
    staff_day = database.query(BiometricAttendanceDay).filter_by(staff_user_id=staff.id).one()
    assert staff_day.student_id is None
    assert staff_day.last_punch_at is not None
    register = database.query(AttendanceRegister).filter_by(register_kind="biometric").one()
    assert database.query(AttendanceEntry).filter_by(register_id=register.id).count() == 1
    staff_rows = client.get("/api/attendance/staff-biometric", headers=headers)
    assert staff_rows.status_code == 200
    assert staff_rows.json()["records"][0]["fullName"] == "Office Staff"
    assert staff_rows.json()["records"][0]["departureAt"] is not None


def test_biometric_import_keeps_unassigned_staff_visible(client, database):
    operator, tatva, _ = setup_students(database)
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    content = (
        b"User ID,Name,Timestamp\n"
        b"101,Tatva Student,2026-08-10 08:15:00\n"
        b"900,Unknown,2026-08-10 13:25:00\n"
    )
    staged = preview(client, headers, content)
    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "Attendance",
        "deviceIdColumn": "User ID",
        "nameColumn": "Name",
        "datetimeColumn": "Timestamp",
        "dateColumn": None,
        "timeColumn": None,
    }
    analysis = client.post(
        "/api/attendance/biometric-imports/analyze",
        headers=headers,
        json=selection,
    )
    people = {item["deviceUserId"]: item for item in analysis.json()["deviceUsers"]}
    assert people["900"]["unassignedStaff"] is False
    committed = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={
            **selection,
            "mappings": [
                {"deviceUserId": "101", "studentId": tatva.id},
                {"deviceUserId": "900", "unassignedStaff": True},
            ],
        },
    )
    assert committed.status_code == 200, committed.text
    assert committed.json()["unassignedStaffDeviceIds"] == ["900"]
    staff_day = database.query(BiometricAttendanceDay).filter_by(device_user_id="900").one()
    assert staff_day.student_id is None
    assert staff_day.staff_user_id is None
    staff_rows = client.get("/api/attendance/staff-biometric", headers=headers)
    assert staff_rows.status_code == 200
    record = staff_rows.json()["records"][0]
    assert record["fullName"] == "Unknown"
    assert record["deviceUserId"] == "900"


@pytest.mark.parametrize("name, expected", [
    ("Vinay Barhate", "Director"),
    (" vinay  barhate ", "Director"),
    ("Dr. Vinay Barhate", "Director"),
    ("DR VINAY BARHATE", "Director"),
    ("Vinay Deshmukh", None),
    ("Pooja Kamble", None),
    ("Vinay Barhate Junior", None),
    ("", None),
])
def test_staff_designation_is_an_exact_display_only_override(name, expected):
    assert _staff_designation(name) == expected


@pytest.mark.parametrize("linked_account", [False, True])
def test_vinay_attendance_displays_director_without_changing_access(client, database, linked_account):
    operator, _, _ = setup_students(database)
    staff = None
    if linked_account:
        staff = User(
            mobile="9000000199",
            full_name="Dr. Vinay Barhate",
            role="front_desk",
            password_hash=hash_password("Password123!"),
        )
        database.add(staff)
        database.commit()
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    mapping = {"staffUserId": staff.id} if staff else {"unassignedStaff": True}
    # A later biometric import must leave the designation and historical rows
    # intact, including when there is no linked application login.
    for day in ("01", "02"):
        content = (
            "User ID,Name,Timestamp\n"
            f"003,Vinay Barhate,2026-09-{day} 12:07:00\n"
            f"003,Vinay Barhate,2026-09-{day} 14:48:00\n"
            f"26,Pooja Kamble,2026-09-{day} 10:34:00\n"
        ).encode()
        staged = preview(client, headers, content)
        committed = client.post(
            "/api/attendance/biometric-imports",
            headers=headers,
            json={
                "previewToken": staged["previewToken"],
                "sheetName": "Attendance",
                "deviceIdColumn": "User ID",
                "nameColumn": "Name",
                "datetimeColumn": "Timestamp",
                "mappings": [
                    {"deviceUserId": "003", **mapping},
                    {"deviceUserId": "26", "unassignedStaff": True},
                ],
            },
        )
        assert committed.status_code == 200, committed.text

    result = client.get("/api/attendance/staff-biometric", headers=headers)
    assert result.status_code == 200
    records = result.json()["records"]
    assert result.json()["staffCount"] == 1
    assert result.json()["directorCount"] == 1
    assert result.json()["recordCount"] == 4
    vinay_rows = [row for row in records if row["deviceUserId"] == "003"]
    assert len(vinay_rows) == 2
    assert {row["date"] for row in vinay_rows} == {"2026-09-01", "2026-09-02"}
    assert all(row["designation"] == "Director" for row in vinay_rows)
    assert all(row["attendanceGroup"] == "directors" for row in vinay_rows)
    assert all(row["role"] == ("front_desk" if staff else "staff") for row in vinay_rows)
    assert all(row["departureAt"] for row in vinay_rows)
    assert all(row["designation"] is None for row in records if row["deviceUserId"] == "26")
    assert all(row["attendanceGroup"] == "staff" for row in records if row["deviceUserId"] == "26")
    identity = database.query(DeviceAttendanceIdentity).filter_by(device_user_id="003").one()
    assert identity.device_name == "Vinay Barhate"
    assert identity.staff_user_id == (staff.id if staff else None)
    assert database.query(AttendanceEntry).count() == 0
    if staff:
        database.refresh(staff)
        assert staff.role == "front_desk"
        assert client.get(
            "/api/attendance/staff-biometric",
            headers={"Authorization": f"Bearer {create_token(staff)}"},
        ).status_code == 403


def test_biometric_import_allows_two_device_ids_for_one_student(client, database):
    operator, tatva, _ = setup_students(database)
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    content = (
        b"User ID,Name,Timestamp\n"
        b"50,Tatva Student,2026-08-10 13:25:00\n"
        b"T-1,Tatva Student,2026-08-11 08:15:00\n"
    )
    staged = preview(client, headers, content)
    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "Attendance",
        "deviceIdColumn": "User ID",
        "nameColumn": "Name",
        "datetimeColumn": "Timestamp",
        "dateColumn": None,
        "timeColumn": None,
    }
    committed = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={
            **selection,
            "mappings": [
                {"deviceUserId": "50", "studentId": tatva.id},
                {"deviceUserId": "T-1", "studentId": tatva.id},
            ],
        },
    )
    assert committed.status_code == 200, committed.text
    identities = database.query(DeviceAttendanceIdentity).filter_by(student_id=tatva.id).all()
    assert {item.device_user_id for item in identities} == {"50", "T-1"}
    days = database.query(BiometricAttendanceDay).filter_by(student_id=tatva.id).all()
    assert {item.attendance_date.isoformat() for item in days} == {"2026-08-10", "2026-08-11"}
    registers = database.query(AttendanceRegister).filter_by(register_kind="biometric").all()
    assert len(registers) == 2
    assert all(
        database.query(AttendanceEntry).filter_by(register_id=register.id, student_id=tatva.id).one().status == "present"
        for register in registers
    )


def test_biometric_import_rejects_unmapped_device_ids_and_duplicate_files(client, database):
    operator, tatva, _ = setup_students(database)
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    content = b"User ID,Timestamp\n101,2026-08-12 08:15:00\n"
    staged = preview(client, headers, content)
    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "Attendance",
        "deviceIdColumn": "User ID",
        "datetimeColumn": "Timestamp",
        "dateColumn": None,
        "timeColumn": None,
        "nameColumn": None,
    }
    unmapped = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={**selection, "mappings": []},
    )
    assert unmapped.status_code == 409

    first = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={**selection, "mappings": [{"deviceUserId": "101", "studentId": tatva.id, "ignore": False}]},
    )
    assert first.status_code == 200
    staged_again = preview(client, headers, content)
    duplicate = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={**selection, "previewToken": staged_again["previewToken"], "mappings": []},
    )
    assert duplicate.status_code == 200
    assert duplicate.json()["alreadyImported"] is True
    assert duplicate.json()["punchesCreated"] == 0
    assert "already imported" in duplicate.json()["message"]
    assert database.query(BiometricImportBatch).count() == 1


def test_parent_cannot_preview_biometric_attendance(client, parent_headers):
    response = client.post(
        "/api/attendance/biometric-imports/preview",
        headers=parent_headers,
        files={"file": ("attendance.csv", b"User ID,Timestamp\n1,2026-08-12 08:00:00\n", "text/csv")},
    )
    assert response.status_code == 403


def test_xlsx_biometric_preview(client, database):
    from openpyxl import Workbook

    operator, _, _ = setup_students(database)
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Punches"
    sheet.append(["Enroll ID", "Punch Time"])
    sheet.append([101, datetime(2026, 8, 12, 8, 15)])
    stream = BytesIO()
    workbook.save(stream)
    staged = preview(client, headers, stream.getvalue(), "attendance.xlsx")
    assert staged["sheets"][0]["name"] == "Punches"
    assert staged["sheets"][0]["detected"]["device_id"] == "Enroll ID"
    assert staged["sheets"][0]["detected"]["datetime"] == "Punch Time"


def test_essl_form_j_excel_preview_and_student_code_suggestion(client, database):
    from openpyxl import Workbook

    operator, tatva, essential = setup_students(database)
    essential_profile = database.query(StudentAcademicProfile).filter_by(student_id=essential.id).one()
    essential_profile.source_student_code = "E-02"
    database.commit()
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DetailedFormJ"
    sheet.append(['FORM "J"'])
    sheet.append(["REGISTER OF EMPLOYMENT"])
    sheet.append(["For The Month Ending August To 2026"])
    sheet.append(["Sr No.", "Employee", "Type", "1 St", "2 S", "3 M", "4 T", "5 W"])
    sheet.append([1, "Name:Device spelling", "M"])
    sheet.append([None, "Code:T-1", "InTime", "08:15", None, "09:05"])
    sheet.append([2, "Name:Vidhisha patil", "F"])
    sheet.append([None, "Code:51", "InTime", "09:10"])
    stream = BytesIO()
    workbook.save(stream)

    staged = preview(client, headers, stream.getvalue(), "form-j.xlsx")
    assert staged["sourceFormat"] == "essl_form_j_workbook"
    assert staged["report"]["reportMonth"] == "2026-08"
    assert staged["report"]["identityCount"] == 2
    assert staged["sheets"][0]["rowCount"] == 3

    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": "DetailedFormJ",
        "deviceIdColumn": "Device Code",
        "nameColumn": "Student Name",
        "datetimeColumn": None,
        "dateColumn": "Date",
        "timeColumn": "InTime",
    }
    analysis = client.post(
        "/api/attendance/biometric-imports/analyze",
        headers=headers,
        json=selection,
    )
    assert analysis.status_code == 200, analysis.text
    people = {item["deviceUserId"]: item for item in analysis.json()["deviceUsers"]}
    assert people["T-1"]["studentId"] == tatva.id
    assert people["T-1"]["matchReason"] == "Student code"
    assert people["T-1"]["dayCount"] == 2
    assert people["51"]["studentId"] == essential.id
    assert people["51"]["matchReason"] == "Confirmed name merge"


def test_essl_form_j_single_day_preview_combines_every_worksheet(client, database):
    from openpyxl import Workbook

    operator, tatva, _ = setup_students(database)
    staff = User(
        mobile="9000000299",
        full_name="Dr. Vinay Barhate",
        role="director",
        password_hash=hash_password("Password123!"),
    )
    database.add(staff)
    database.commit()
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    workbook = Workbook()
    student_sheet = workbook.active
    student_sheet.title = "Sheet1"
    staff_sheet = workbook.create_sheet("Sheet2")

    for sheet in (student_sheet, staff_sheet):
        sheet.append(['FORM "J"'])
        sheet.append(["REGISTER OF EMPLOYMENT"])
        sheet.append(["For The Month Ending September To 2026"])
        sheet.append(["Sr No.", "Employee", "Type", "1 T"])
    student_sheet.append([1, "Name:Tatva Student", "M"])
    student_sheet.append([None, "Code:T-1", "InTime", "08:15"])
    staff_sheet.append([1, "Name:Vinay Barhate", "M"])
    staff_sheet.append([None, "Code:003", "InTime", "13:44"])
    staff_sheet.append([2, "Name:Pooja Kamble", "F"])
    staff_sheet.append([None, "Code:26", "InTime", "09:10"])

    stream = BytesIO()
    workbook.save(stream)
    staged = preview(client, headers, stream.getvalue(), "one-day-form-j.xlsx")

    assert staged["sourceFormat"] == "essl_form_j_workbook"
    assert staged["report"]["reportMonth"] == "2026-09"
    assert staged["report"]["sourceSheets"] == ["Sheet1", "Sheet2"]
    assert staged["report"]["staffSourceSheets"] == ["Sheet2"]
    assert staged["report"]["identityCount"] == 3
    assert staged["sheets"][0]["name"] == "All Form J worksheets"
    assert staged["sheets"][0]["rowCount"] == 3

    selection = {
        "previewToken": staged["previewToken"],
        "sheetName": staged["sheets"][0]["name"],
        "deviceIdColumn": "Device Code",
        "nameColumn": "Student Name",
        "datetimeColumn": None,
        "dateColumn": "Date",
        "timeColumn": "InTime",
    }
    analysis = client.post(
        "/api/attendance/biometric-imports/analyze",
        headers=headers,
        json=selection,
    )
    assert analysis.status_code == 200, analysis.text
    result = analysis.json()
    assert result["rowsSeen"] == 3
    assert result["uniqueAttendanceDays"] == 3
    people = {item["deviceUserId"]: item for item in result["deviceUsers"]}
    assert people["T-1"]["studentId"] == tatva.id
    assert people["003"]["staffUserId"] == staff.id
    assert people["003"]["matchReason"] == "Sheet2 · staff name"
    assert people["26"]["unassignedStaff"] is True
    assert people["26"]["matchReason"] == "Sheet2 · staff attendance"

    committed = client.post(
        "/api/attendance/biometric-imports",
        headers=headers,
        json={
            **selection,
            "mappings": [
                {"deviceUserId": "T-1", "studentId": tatva.id},
                {"deviceUserId": "003", "staffUserId": staff.id},
                {"deviceUserId": "26", "unassignedStaff": True},
            ],
        },
    )
    assert committed.status_code == 200, committed.text
    assert committed.json()["staffPunchesCreated"] == 2
    assert committed.json()["matchedStaff"] == 1
    staff_rows = client.get("/api/attendance/staff-biometric", headers=headers)
    assert staff_rows.status_code == 200
    names = {item["fullName"] for item in staff_rows.json()["records"]}
    assert names == {"Dr. Vinay Barhate", "Pooja Kamble"}
    assert staff_rows.json()["staffCount"] == 1
    assert staff_rows.json()["directorCount"] == 1


def test_essl_form_j_sheet_parser_keeps_first_punches():
    sheet = SheetData("DetailedFormJ", [
        ['FORM "J"'],
        ["REGISTER OF EMPLOYMENT"],
        ["For The Month Ending August To 2026"],
        ["Sr No.", "Employee", "Type", "1 St", "2 S", "3 M", "4 T", "5 W"],
        [1, "Name:Kamal Parsatwar", "M"],
        [None, "Code:T-1", "InTime", "09:59", None, "11:32"],
        [None, "Designation:", "OutTime", "14:59", None, "13:43"],
        [None, "DOJ:01-Aug-2026", "Status", "P", "A", "P"],
    ])
    punches, errors, report = parse_essl_form_j_sheet(sheet)
    assert errors == []
    assert report["identityCount"] == 1
    assert report["reportMonth"] == "2026-08"
    assert [(item["attendanceDate"].day, item["firstPunchAt"].hour, item["firstPunchAt"].minute) for item in punches] == [
        (1, 4, 29),
        (3, 6, 2),
    ]


def test_essl_monthly_work_duration_parser_keeps_daily_and_monthly_hours():
    sheet = SheetData("WorkDurationReportFourPunch", [
        [None, "Monthly Status Report (Detailed Work Duration(Four Punch))"],
        [None, "Aug 01 2026  To  Aug 31 2026"],
        ["Days", None, "1 St", "2 S", "3 M"],
        ["Employee:", None, None, None, "7 : Sneha", None, None, None,
         " Total Work Duration: 7:30 Hrs. Total OT: 8:30 Hrs. Present: 2 Absent: 1 WeeklyOff: 1 Holidays: 0 Leaves Taken: 0 Late By Hrs: 00:00 Late By Days: 0 Early By Hrs: 00:00 Early going By Days: 0 Total Duration(+OT): 16:00 Average Working Hrs: 8:00"],
        ["Status", None, "WOP", "A", "P"],
        ["InTime1", None, "10:00", None, "10:30"],
        ["OutTime1", None, "18:30", None, "18:00"],
        ["InTime2"],
        ["OutTime2"],
        ["Duration", None, "00:00", "00:00", "7:30"],
        ["OT", None, "8:30", "00:00", "00:00"],
    ])
    punches, workdays, errors, report = parse_essl_work_duration_sheet(sheet)
    assert errors == []
    assert report["reportMonth"] == "2026-08"
    assert report["identities"][0]["totalWorkMinutes"] == 960
    assert len(punches) == 2
    assert [item["attendanceStatus"] for item in workdays] == ["weekly_off_present", "absent", "present"]
    assert [item["workDurationMinutes"] for item in workdays] == [510, 0, 450]


def test_monthly_work_duration_upload_publishes_staff_daily_hours(client, database):
    from openpyxl import Workbook

    operator, _, _ = setup_students(database)
    staff = User(
        mobile="9000000399", full_name="Sneha", role="front_desk",
        password_hash=hash_password("Password123!"),
    )
    database.add(staff)
    database.commit()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WorkDurationReportFourPunch"
    for row in [
        [None, "Monthly Status Report (Detailed Work Duration(Four Punch))"],
        [None, "Aug 01 2026  To  Aug 31 2026"],
        ["Days", None, "1 St", "2 S", "3 M"],
        ["Employee:", None, None, None, "7 : Sneha", None, None, None,
         " Total Work Duration: 7:30 Hrs. Total OT: 8:30 Hrs. Present: 2 Absent: 1 WeeklyOff: 1 Holidays: 0 Leaves Taken: 0 Late By Hrs: 00:00 Late By Days: 0 Early By Hrs: 00:00 Early going By Days: 0 Total Duration(+OT): 16:00 Average Working Hrs: 8:00"],
        ["Status", None, "WOP", "A", "P"],
        ["InTime1", None, "10:00", None, "10:30"],
        ["OutTime1", None, "18:30", None, "18:00"],
        ["Duration", None, "00:00", "00:00", "7:30"],
        ["OT", None, "8:30", "00:00", "00:00"],
    ]:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    headers = {"Authorization": f"Bearer {create_token(operator)}"}
    staged = preview(client, headers, stream.getvalue(), "staff-duration.xlsx")
    assert staged["sourceFormat"] == "essl_work_duration"
    selection = {
        "previewToken": staged["previewToken"], "sheetName": sheet.title,
        "deviceIdColumn": "Device Code", "nameColumn": "Staff Name",
        "datetimeColumn": None, "dateColumn": "Date", "timeColumn": "Work Time",
    }
    analysis = client.post("/api/attendance/biometric-imports/analyze", headers=headers, json=selection)
    assert analysis.status_code == 200, analysis.text
    assert analysis.json()["deviceUsers"][0]["staffUserId"] == staff.id
    committed = client.post("/api/attendance/biometric-imports", headers=headers, json={
        **selection, "mappings": [{"deviceUserId": "7", "staffUserId": staff.id}],
    })
    assert committed.status_code == 200, committed.text
    assert committed.json()["staffWorkdaysCreated"] == 3
    assert database.query(StaffAttendanceWorkday).count() == 3
    register = client.get("/api/attendance/staff-biometric", headers=headers).json()
    assert len(register["records"]) == 3
    absent = next(item for item in register["records"] if item["attendanceStatus"] == "absent")
    assert absent["arrivalAt"] is None
    assert register["monthlyTotals"][0]["totalWorkMinutes"] == 960


def test_essl_detailed_form_j_without_month_uses_upload_month():
    sheet = SheetData("DetailedFormJ", [
        [None, 'FORM "J"'],
        [None, "REGISTER OF EMPLOYMENT"],
        ["Sr No.", "Employee", "Type", "1 St", "2 S", "3 M", "20 Th", "23 S"],
        [1, "Name:Kamal Parsatwar", "M"],
        [None, "Code:T-1", "InTime", None, None, None, "09:59", "11:32"],
        [None, "Designation:", "OutTime", None, None, None, "14:59", "13:43"],
    ])
    punches, errors, report = parse_essl_form_j_sheet(
        sheet,
        reference_date=date(2026, 8, 28),
    )
    assert errors == []
    assert report["reportMonth"] == "2026-08"
    assert report["monthSource"] == "upload_date"
    assert [item["attendanceDate"].isoformat() for item in punches] == [
        "2026-08-20",
        "2026-08-23",
    ]


def test_essl_form_j_pdf_parser(monkeypatch):
    class FakePage:
        def extract_text(self):
            return 'FORM "J"\nGenerated By:essl\nFor The Month Ending August To 2026'

        def extract_words(self, **_kwargs):
            return [
                {"text": "Name:Vidhisha", "x0": 27.2, "x1": 81.4, "top": 185.2},
                {"text": "patil", "x0": 27.2, "x1": 45.0, "top": 194.4},
                {"text": "Code:51", "x0": 27.2, "x1": 57.4, "top": 203.2},
                {"text": "InTime", "x0": 166.0, "x1": 190.4, "top": 204.2},
                {"text": "14:06", "x0": 493.2, "x1": 513.2, "top": 203.2},
                {"text": "10:22", "x0": 518.4, "x1": 538.4, "top": 203.2},
            ]

    class FakePdf:
        pages = [FakePage()]

        def close(self):
            pass

    import pdfplumber
    monkeypatch.setattr(pdfplumber, "open", lambda _stream: FakePdf())
    punches, errors, metadata = parse_essl_form_j_pdf(b"fake-pdf")
    assert errors == []
    assert metadata["format"] == "essl_form_j"
    assert metadata["reportMonth"] == "2026-08"
    assert metadata["identityCount"] == 1
    assert metadata["identities"] == [{"deviceUserId": "51", "deviceName": "Vidhisha patil"}]
    assert [(item["attendanceDate"].day, item["firstPunchAt"].hour, item["firstPunchAt"].minute) for item in punches] == [
        (11, 8, 36),
        (12, 4, 52),
    ]
