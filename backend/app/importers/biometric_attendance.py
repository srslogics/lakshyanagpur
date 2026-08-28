from __future__ import annotations

import csv
import calendar
import io
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


INDIA_TZ = ZoneInfo("Asia/Kolkata")
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 100_000
SUPPORTED_EXTENSIONS = {"csv", "txt", "xlsx", "xls"}
SUPPORTED_UPLOAD_EXTENSIONS = SUPPORTED_EXTENSIONS | {"pdf"}


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: list[list[Any]]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).lower())


def normalize_device_id(value: Any) -> str:
    text = _text(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _decode_delimited(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("The text file encoding is not supported")


def _delimited_rows(content: bytes) -> list[list[str]]:
    text = _decode_delimited(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel
    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return [row for row in rows if any(_text(cell) for cell in row)]


def read_workbook(content: bytes, filename: str) -> list[SheetData]:
    if not content:
        raise ValueError("The selected attendance file is empty")
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError("Attendance files must be 5 MB or smaller")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Upload a CSV, TXT, XLS or XLSX attendance export")
    if extension in {"csv", "txt"}:
        sheets = [SheetData("Attendance", _delimited_rows(content))]
    elif extension == "xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise ValueError("The XLSX attendance file could not be read") from error
        sheets = [
            SheetData(sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
    else:
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception as error:
            raise ValueError("The XLS attendance file could not be read") from error
        sheets = []
        for name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(name)
            rows = []
            for row_index in range(worksheet.nrows):
                row = []
                for cell in worksheet.row(row_index):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(value, workbook.datemode)
                    row.append(value)
                rows.append(row)
            sheets.append(SheetData(name, rows))
    cleaned = []
    total_rows = 0
    for sheet in sheets:
        rows = [row for row in sheet.rows if any(_text(cell) for cell in row)]
        total_rows += len(rows)
        cleaned.append(SheetData(sheet.name, rows))
    if total_rows > MAX_IMPORT_ROWS:
        raise ValueError("The attendance export contains more than 100,000 rows")
    if not any(sheet.rows for sheet in cleaned):
        raise ValueError("No attendance rows were found in the selected file")
    return cleaned


HEADER_ALIASES = {
    "device_id": {
        "userid", "useridno", "userno", "enrollid", "enrollmentid", "enrolid",
        "empcode", "employeecode", "employeeid", "acno", "pin", "id", "code",
    },
    "name": {"name", "username", "employeename", "studentname", "personname"},
    "date": {"date", "attendancedate", "punchdate", "logdate", "recorddate"},
    "time": {"time", "punchtime", "logtime", "checktime", "recordtime"},
    "datetime": {
        "datetime", "datetimerecord", "timestamp", "punchdatetime", "checkdatetime",
        "attendancetime", "logdatetime", "recorddatetime",
    },
}


def detect_columns(headers: list[Any]) -> dict[str, str | None]:
    normalized = {normalize_header(header): _text(header) for header in headers if _text(header)}
    result: dict[str, str | None] = {key: None for key in HEADER_ALIASES}
    for target, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[target] = normalized[alias]
                break
    if result["datetime"] and result["date"] and result["datetime"] == result["date"]:
        result["date"] = None
    if result["datetime"] and result["time"] and result["datetime"] == result["time"]:
        result["time"] = None
    return result


def sheet_preview(sheet: SheetData) -> dict:
    if not sheet.rows:
        return {"name": sheet.name, "headers": [], "rows": [], "detected": {}}
    headers = [_text(value) or f"Column {index + 1}" for index, value in enumerate(sheet.rows[0])]
    width = len(headers)
    preview_rows = []
    for row in sheet.rows[1:6]:
        preview_rows.append({headers[index]: _text(row[index]) if index < len(row) else "" for index in range(width)})
    detected = detect_columns(headers)
    # Many biometric exports call a combined date-time field "Punch Time".
    # Treat it as a timestamp when its actual worksheet value includes a date;
    # preserve it as a time-only field when a separate Date column exists.
    time_header = detected.get("time")
    if time_header and not detected.get("date") and not detected.get("datetime"):
        position = headers.index(time_header)
        sample_values = [
            row[position]
            for row in sheet.rows[1:6]
            if position < len(row) and row[position] not in (None, "")
        ]
        has_combined_value = any(
            isinstance(value, datetime)
            or bool(re.search(r"\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}", _text(value)))
            for value in sample_values
        )
        if has_combined_value:
            detected["datetime"] = time_header
            detected["time"] = None
    return {
        "name": sheet.name,
        "headers": headers,
        "rows": preview_rows,
        "rowCount": max(0, len(sheet.rows) - 1),
        "detected": detected,
    }


DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y",
    "%d-%b-%Y", "%d %b %Y", "%d.%m.%Y",
)
TIME_FORMATS = ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p")
DATETIME_FORMATS = tuple(
    f"{date_format} {time_format}"
    for date_format in DATE_FORMATS
    for time_format in TIME_FORMATS
) + (
    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
)


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for format_string in DATE_FORMATS:
        try:
            return datetime.strptime(text, format_string).date()
        except ValueError:
            pass
    raise ValueError(f"Unrecognised attendance date: {text or 'blank'}")


def _parse_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        total_seconds = round(float(value) * 86400)
        return time((total_seconds // 3600) % 24, (total_seconds % 3600) // 60, total_seconds % 60)
    text = _text(value).upper()
    for format_string in TIME_FORMATS:
        try:
            return datetime.strptime(text, format_string).time()
        except ValueError:
            pass
    raise ValueError(f"Unrecognised punch time: {text or 'blank'}")


def parse_datetime(datetime_value: Any = None, date_value: Any = None, time_value: Any = None) -> datetime:
    if isinstance(datetime_value, datetime):
        parsed = datetime_value
    elif isinstance(datetime_value, date):
        parsed = datetime.combine(datetime_value, time.min)
    elif _text(datetime_value):
        text = _text(datetime_value).replace("  ", " ")
        parsed = None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            for format_string in DATETIME_FORMATS:
                try:
                    parsed = datetime.strptime(text.upper(), format_string)
                    break
                except ValueError:
                    pass
        if parsed is None:
            raise ValueError(f"Unrecognised punch date and time: {text}")
    else:
        parsed = datetime.combine(_parse_date(date_value), _parse_time(time_value))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=INDIA_TZ)
    return parsed.astimezone(timezone.utc)


def parse_punches(
    sheet: SheetData,
    device_id_column: str,
    datetime_column: str | None = None,
    date_column: str | None = None,
    time_column: str | None = None,
    name_column: str | None = None,
) -> tuple[list[dict], list[dict]]:
    if not sheet.rows:
        return [], []
    headers = [_text(value) or f"Column {index + 1}" for index, value in enumerate(sheet.rows[0])]
    positions = {header: index for index, header in enumerate(headers)}
    required = [device_id_column]
    if datetime_column:
        required.append(datetime_column)
    else:
        required.extend([date_column, time_column])
    missing = [column for column in required if not column or column not in positions]
    if missing:
        raise ValueError("Choose the device ID and punch date/time columns")

    daily_punches: dict[tuple[str, date], dict] = {}
    errors = []
    raw_rows = 0
    for source_row, row in enumerate(sheet.rows[1:], start=2):
        values = {header: row[index] if index < len(row) else None for header, index in positions.items()}
        if not any(_text(value) for value in values.values()):
            continue
        raw_rows += 1
        device_user_id = normalize_device_id(values.get(device_id_column))
        if not device_user_id:
            errors.append({"row": source_row, "message": "Device ID is blank"})
            continue
        try:
            punch_at = parse_datetime(
                values.get(datetime_column) if datetime_column else None,
                values.get(date_column) if date_column else None,
                values.get(time_column) if time_column else None,
            )
        except ValueError as error:
            errors.append({"row": source_row, "message": str(error)})
            continue
        local_day = punch_at.astimezone(INDIA_TZ).date()
        item = {
            "deviceUserId": device_user_id,
            "deviceName": _text(values.get(name_column)) if name_column else "",
            "attendanceDate": local_day,
            "firstPunchAt": punch_at,
            "lastPunchAt": None,
            "sourceRow": source_row,
        }
        key = (device_user_id, local_day)
        existing = daily_punches.get(key)
        if not existing:
            daily_punches[key] = item
        else:
            all_punches = [existing["firstPunchAt"], punch_at]
            if existing.get("lastPunchAt"):
                all_punches.append(existing["lastPunchAt"])
            existing["firstPunchAt"] = min(all_punches)
            existing["lastPunchAt"] = max(all_punches)
            if not existing["deviceName"] and item["deviceName"]:
                existing["deviceName"] = item["deviceName"]
    rows = sorted(daily_punches.values(), key=lambda item: (item["attendanceDate"], item["deviceUserId"]))
    for item in rows:
        item["rowsSeen"] = raw_rows
    return rows, errors


def _essl_report_month(report_text: str) -> tuple[int, int]:
    month_match = re.search(
        r"For\s+The\s+Month\s+Ending\s+([A-Za-z]+)\s+(?:To\s+)?(\d{4})",
        report_text,
        re.IGNORECASE,
    )
    if not month_match:
        raise ValueError("The attendance month could not be read from this report")
    try:
        month = list(calendar.month_name).index(month_match.group(1).title())
        year = int(month_match.group(2))
    except (ValueError, IndexError) as error:
        raise ValueError("The attendance month in this report is not valid") from error
    return year, month


def is_essl_form_j_sheet(sheet: SheetData) -> bool:
    report_text = "\n".join(_text(cell) for row in sheet.rows for cell in row if _text(cell))
    upper = report_text.upper()
    has_form_heading = 'FORM "J"' in upper and "REGISTER OF EMPLOYMENT" in upper
    has_punch_layout = (
        "INTIME" in upper
        and "OUTTIME" in upper
        and bool(re.search(r"(?:^|\n)CODE\s*:[^\n]+", upper))
    )
    # Some eSSL DetailedFormJ exports omit the printed month caption even though
    # the day columns and biometric rows are complete. They are still Form J
    # workbooks and must not fall through to the generic column-mapping screen.
    return has_form_heading and ("FOR THE MONTH ENDING" in upper or has_punch_layout)


def _form_j_day_columns(sheet: SheetData, month_days: int = 31) -> dict[int, int]:
    day_columns: dict[int, int] = {}
    for row in sheet.rows:
        detected_days = {}
        for column, value in enumerate(row):
            match = re.match(r"^\s*(\d{1,2})(?:\s|$)", _text(value))
            if match and 1 <= int(match.group(1)) <= month_days:
                detected_days[column] = int(match.group(1))
        if len(detected_days) >= 5:
            day_columns = detected_days
    return day_columns


def _form_j_active_days(sheet: SheetData, day_columns: dict[int, int]) -> set[int]:
    active_days: set[int] = set()
    for row in sheet.rows:
        if not any(_text(value) == "InTime" for value in row):
            continue
        for column, day in day_columns.items():
            value = row[column] if column < len(row) else None
            if _text(value) and _text(value) not in {"0", "00:00", "00:00:00"}:
                active_days.add(day)
    return active_days


def _form_j_report_month(
    report_text: str,
    active_days: set[int],
    reference_date: date | None = None,
) -> tuple[int, int, str]:
    try:
        year, month = _essl_report_month(report_text)
        return year, month, "report"
    except ValueError:
        # eSSL sometimes leaves the month caption blank in DetailedFormJ XLS
        # exports. Daily uploads are for the current month; immediately after a
        # month boundary, punch days greater than today's date identify the
        # previous month. Exposing the source in metadata keeps this auditable.
        reference = reference_date or datetime.now(INDIA_TZ).date()
        if active_days and max(active_days) > reference.day:
            if reference.month == 1:
                return reference.year - 1, 12, "upload_date"
            return reference.year, reference.month - 1, "upload_date"
        return reference.year, reference.month, "upload_date"


def parse_essl_form_j_sheet(
    sheet: SheetData,
    reference_date: date | None = None,
    report_month: str | None = None,
) -> tuple[list[dict], list[dict], dict]:
    """Read an eSSL Form J XLS/XLSX worksheet and retain each first daily punch."""
    if not is_essl_form_j_sheet(sheet):
        raise ValueError("This worksheet is not an eSSL Form J attendance report")
    report_text = "\n".join(_text(cell) for row in sheet.rows for cell in row if _text(cell))
    day_columns = _form_j_day_columns(sheet)
    active_days = _form_j_active_days(sheet, day_columns)
    if report_month:
        try:
            year, month = (int(value) for value in report_month.split("-", 1))
            if not 1 <= month <= 12:
                raise ValueError
        except (TypeError, ValueError) as error:
            raise ValueError("The saved attendance month is not valid") from error
        month_source = "preview"
    else:
        year, month, month_source = _form_j_report_month(report_text, active_days, reference_date)
    month_days = calendar.monthrange(year, month)[1]
    day_columns = {column: day for column, day in day_columns.items() if day <= month_days}
    current_name = ""
    identities: dict[str, str] = {}
    daily_punches: dict[tuple[str, date], dict] = {}
    errors = []
    rows_seen = 0

    for row_index, row in enumerate(sheet.rows):
        source_row = row_index + 1
        for value in row:
            text = _text(value)
            if text.startswith("Name:"):
                current_name = text[5:].strip()

        code_cell = next((_text(value) for value in row if _text(value).startswith("Code:")), "")
        if not code_cell:
            continue
        device_user_id = normalize_device_id(code_cell[5:])
        if not device_user_id:
            errors.append({"row": source_row, "message": "A biometric code is blank"})
            continue
        identities.setdefault(device_user_id, current_name)
        if not day_columns:
            errors.append({"row": source_row, "message": "The Form J day columns could not be read"})
            continue
        out_row = sheet.rows[row_index + 1] if row_index + 1 < len(sheet.rows) else []
        is_out_row = any(_text(value) == "OutTime" for value in out_row)
        for column, day in day_columns.items():
            value = row[column] if column < len(row) else None
            text = _text(value)
            if not text or text in {"0", "00:00", "00:00:00"}:
                continue
            rows_seen += 1
            try:
                punch_at = parse_datetime(date_value=date(year, month, day), time_value=value)
            except ValueError as error:
                errors.append({"row": source_row, "message": str(error)})
                continue
            item = {
                "deviceUserId": device_user_id,
                "deviceName": current_name,
                "attendanceDate": date(year, month, day),
                "firstPunchAt": punch_at,
                "lastPunchAt": None,
                "sourceRow": source_row,
            }
            if is_out_row:
                out_value = out_row[column] if column < len(out_row) else None
                if _text(out_value) and _text(out_value) not in {"0", "00:00", "00:00:00"}:
                    try:
                        out_at = parse_datetime(date_value=date(year, month, day), time_value=out_value)
                        if out_at >= punch_at:
                            item["lastPunchAt"] = out_at
                    except ValueError as error:
                        errors.append({"row": source_row + 1, "message": str(error)})
            key = (device_user_id, item["attendanceDate"])
            existing = daily_punches.get(key)
            if not existing:
                daily_punches[key] = item
            else:
                existing["firstPunchAt"] = min(existing["firstPunchAt"], item["firstPunchAt"])
                last_values = [value for value in (existing.get("lastPunchAt"), item.get("lastPunchAt")) if value]
                existing["lastPunchAt"] = max(last_values) if last_values else None

    if not identities:
        raise ValueError("No biometric identities were found in this worksheet")
    punches = sorted(daily_punches.values(), key=lambda item: (item["attendanceDate"], item["deviceUserId"]))
    for item in punches:
        item["rowsSeen"] = rows_seen
    return punches, errors, {
        "format": "essl_form_j_workbook",
        "reportMonth": f"{year:04d}-{month:02d}",
        "monthSource": month_source,
        "identityCount": len(identities),
        "identities": [
            {"deviceUserId": device_user_id, "deviceName": device_name}
            for device_user_id, device_name in identities.items()
        ],
    }


def parse_essl_form_j_pdf(content: bytes) -> tuple[list[dict], list[dict], dict]:
    """Read the eSSL Form J month register and retain every printed InTime."""
    try:
        import pdfplumber
        workbook = pdfplumber.open(io.BytesIO(content))
    except Exception as error:
        raise ValueError("The PDF attendance report could not be read") from error
    try:
        page_words = [
            page.extract_words(x_tolerance=1, y_tolerance=1, keep_blank_chars=False)
            for page in workbook.pages
        ]
        report_text = "\n".join(page.extract_text() or "" for page in workbook.pages)
    finally:
        workbook.close()
    if 'FORM "J"' not in report_text.upper() or "GENERATED BY:ESSL" not in report_text.upper():
        raise ValueError("This PDF is not an eSSL Form J attendance report")
    year, month = _essl_report_month(report_text)
    month_days = calendar.monthrange(year, month)[1]
    # The report renders one 25.2pt-wide calendar column per date, starting at x=241.2.
    first_day_center = 253.8
    day_width = 25.2
    punches = []
    identities: dict[str, str] = {}
    errors = []
    for page_number, words in enumerate(page_words, start=1):
        ordered = sorted(words, key=lambda item: (item["top"], item["x0"]))
        last_name_index = None
        for index, word in enumerate(ordered):
            text = word["text"].strip()
            if text.startswith("Name:"):
                last_name_index = index
                continue
            if not text.startswith("Code:"):
                continue
            device_user_id = normalize_device_id(text[5:])
            if not device_user_id:
                errors.append({"page": page_number, "message": "A biometric code is blank"})
                continue
            device_name = ""
            if last_name_index is not None:
                name_word = ordered[last_name_index]
                name_parts = [name_word["text"][5:]]
                for item in ordered[last_name_index + 1:index]:
                    candidate = item["text"].strip()
                    if (
                        24 <= item["x0"] < 113
                        and not candidate.startswith(("Code:", "DOJ:", "Designation:", "Name:"))
                        and "Generated By:" not in candidate
                        and candidate not in {"NAME", "OF", "EMPLOYER", "REGISTRATION", "NO"}
                    ):
                        name_parts.append(candidate)
                device_name = " ".join(part for part in name_parts if part).strip()
            identities.setdefault(device_user_id, device_name)
            time_words = [
                item for item in ordered
                if abs(item["top"] - word["top"]) <= 2
                and item["x0"] >= 235
                and re.fullmatch(r"\d{1,2}:\d{2}", item["text"])
            ]
            for item in time_words:
                center = (item["x0"] + item["x1"]) / 2
                day = round((center - first_day_center) / day_width) + 1
                if not 1 <= day <= month_days:
                    continue
                try:
                    punch_at = parse_datetime(
                        date_value=date(year, month, day),
                        time_value=item["text"],
                    )
                except ValueError as error:
                    errors.append({"page": page_number, "message": str(error)})
                    continue
                punches.append({
                    "deviceUserId": device_user_id,
                    "deviceName": device_name,
                    "attendanceDate": date(year, month, day),
                    "firstPunchAt": punch_at,
                    "lastPunchAt": None,
                    "sourcePage": page_number,
                })
    if not identities:
        raise ValueError("No biometric identities were found in this PDF")
    punches.sort(key=lambda item: (item["attendanceDate"], item["deviceUserId"]))
    rows_seen = len(punches)
    for item in punches:
        item["rowsSeen"] = rows_seen
    return punches, errors, {
        "format": "essl_form_j",
        "reportMonth": f"{year:04d}-{month:02d}",
        "identityCount": len(identities),
        "identities": [
            {"deviceUserId": device_user_id, "deviceName": device_name}
            for device_user_id, device_name in identities.items()
        ],
    }
